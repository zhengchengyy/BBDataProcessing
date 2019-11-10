from openpyxl import load_workbook
from openpyxl import Workbook
import time, datetime
from openpyxl.styles import Alignment

inputfilename = 'test20.xlsx'
outputfilename = 'result.xlsx'

workbook = load_workbook(inputfilename)
sheets = workbook.sheetnames  # 从名称获取sheet
booksheet = workbook[sheets[0]] #取第一张表

times = []
amps = []

for row in range(2,booksheet.max_row+1):
	t = booksheet.cell(row, 4).value
	times.append(t)

	r = booksheet.cell(row, 16).value
	amps.append(r)

# print(time)
# print(amp)

from feature_modules import *
from feature_extractor import FeatureExtractor
import time

from matplotlib import pyplot as plt
from matplotlib import style
import numpy as np



feature_names = ["RangeModule", "SDModule", "MeanModule"]


def draw_features_from_db(times, amps):
    title = "features"
    fig = plt.figure(title, figsize=(6, 8))

    # 根据时间采集数据，基本单位为s，比如1s、10s、30s、60s
    # interval表示每次分析的时间跨度，rate表示间隔多长时间进行一次分析
    interval = 1
    rate = 1
    fig.suptitle(" (" + "interval:" + str(interval) + "s, " + "stepsize:" + str(rate) + "s)")

    # 定义特征提取器
    extractor = FeatureExtractor()

    for feature in feature_names:
        # 定义特征提取模块
        module = eval(feature + "(" + str(interval) + "," + str(rate) + ")")
        # 注册特征提取模块
        extractor.register(module)

    # 定义画布左右位置的计数：标签累加，即人数累加
    ntags = 1
    tag_acc = 1

    # read the data that is of a certain action one by one
    for tag in range(ntags):
        inittime, termtime = times[0], times[-1]

        # get the arrays according to which we will plot later
        times = times
        volts = amps

        # 定义存储时间、特征列表
        feature_times, feature_values = {}, {}

        feature_times = []
        from collections import defaultdict
        feature_values = defaultdict(list)
        for feature in feature_names:
            feature_values[feature[:-6]] = []

        # 对每个采集设备进行特征提取 ndevices
        for j in range(len(volts)):
            value = {
                "time": times[j],
                "volt": volts[j]
            }
            # print(times[j])
            output = extractor.process(value)
            if (output):
                features = {
                    "feature_time":times[j],
                    "feature_value": output,
                    "interval": interval,
                    "rate": rate
                }
                feature_times.append(features['feature_time'])
                for feature_type in feature_values.keys():
                    feature_values[feature_type].append(features['feature_value'][feature_type])

        # 清理所有模块，防止过期数据
        extractor.clear()

        # 定义特征数量
        nfeatures = len(feature_values)
        # 定义画布上下位置的计数，即特征累加
        fea_acc = 0
        style.use('default')
        colors = ['r', 'b', 'g', 'c', 'm']  # m c

        for feature_type in feature_values.keys():
            tag_acc += 1
            fea_acc += 1
            ax = fig.add_subplot(nfeatures, ntags, fea_acc)

            plt.subplots_adjust(hspace=0.5)  # 函数中的wspace是子图之间的垂直间距，hspace是子图的上下间距
            ax.set_title(feature_type)

            # ax.set_xlim(feature_times[0], feature_times[-1])
            ax.plot(feature_times, feature_values[feature_type], color='r', label='device', alpha=0.9)

            # 设置每个数据对应的图像名称
            if fea_acc == 1 and tag_acc == 1:
                ax.legend(loc='upper right')
                ax.set_xlabel('Time(s)')
            if fea_acc == nfeatures:
                # 设置人员
                person = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
                ax.set_xlabel("Person" + person[tag_acc - 1] + ": " + str(inittime)
                              + " ~ " + str(termtime))

            # 显示网格
            # ax.grid(True, which='both')


    # 最大化显示图像窗口
    plt.get_current_fig_manager().window.showMaximized()
    plt.show()


draw_features_from_db(times,amps)
